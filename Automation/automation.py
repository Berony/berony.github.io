import openpyxl
import hashlib
import pandas as pd
#import modin.pandas as pd
from typing import Final



import time

#def 로 정의 되는 Const 값
MAX_HEADER_LENGTH : Final[int] = 100
TARGET_FILE : Final[str] = "WIPS_Sample.csv"
TARGET_SHEET_INDEX : Final[int]= 0

#Hash Class
class Hash256():
    def __init__(self):
        self.sha256BOT = hashlib.sha256()

    def hash(self, hashTarget):
        print(hashTarget)
        print(self.sha256BOT)
        self.sha256BOT.update(str(hashTarget).encode())
        return self.sha256BOT.hexdigest



#hashBOT 객체생성
hashBOT = Hash256()

#Header 로드
headerNames = []

#Load CSV
beforeLoad = time.time()
print('before Load...')

df = pd.read_csv(TARGET_FILE)
afterLoad = time.time()
print(f"Pandas Read CSV : {afterLoad - beforeLoad} Lapsed")

# print(df)


#########################
#      Plotly // html로 저장가능
#########################

import plotly.express as px
df = px.data.gapminder().query("continent == 'Oceania'")
fig = px.line(df, x='year', y='lifeExp', color='country', markers=True)
#fig.show(id='the_graph', config= {'displaylogo': False})

fig.show(id='the_graph', config= {'displayModeBar': False})
fig.write_html("test.html", config= {'displayModeBar': False})



#########################
#      Matplotlib
#########################

# import matplotlib.pyplot as plt
# import numpy as np

# x = np.linspace(0, 10, 100)
# y = np.sin(x)

# plt.plot(x, y)
# plt.xlabel('x axis')
# plt.ylabel('y axis')
# plt.title('Sine Wave')
# plt.show()

#########################
#      bokeh
#########################

# from bokeh.plotting import figure, show, save, output_file

# # 데이터 준비
# x = [1, 2, 3, 4, 5]
# y = [6, 7, 2, 4, 5]

# # 빈 캔버스 생성
# p = figure(title="simple line example", x_axis_label='x', y_axis_label='y')

# p.toolbar.logo = None
# p.toolbar_location = None

# # 라인 그래프 추가
# p.line(x, y, legend_label="Temp.", line_width=2)

# # 결과 보여주기
# show(p)

# output_file("test.html")
# save(p)




# i = 1 # start with column '1'
# while (ws.cell(1,i).value != None and i<MAX_HEADER_LENGTH()):
#     headerNames.append(ws.cell(1,i).value)
#     i += 1

# print(headerNames)
# print(headerNames[1])




# if __name__ == '__main__':
#     main()



# for i in range(1,headerLength+1):

#     print(i)
#     print(wb.sheetnames[i])





# for names in wb.sheetnames:
#     print(names)


# ws1 = wb["Sheet"]
# print(ws1.cell(1,1).value)

# ws2 = wb["hi"]
# print(ws2.cell(1,1).value)




#from openpyxl import Workbook


# wb = openpyxl.Workbook()
#wb = Workbook()

# ws = wb.active

# ws["A1"] = "Hello Excel"
# wb.save("hello.xlsx")
